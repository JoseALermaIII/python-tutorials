#! python3
"""Send dues reminders

Sends emails based on payment status in spreadsheet.

Notes:
    * ``smtp_info`` file has each item on a separate line.
    * Email address used is specially created for this chapter.
    * Use :func:`input` or ``sys.argv[1]`` for password to prevent storing in unencrypted file.

"""


def main():
    import openpyxl, smtplib, datetime
    
    # Open the spreadsheet and get the latest dues status.
    wb = openpyxl.load_workbook('duesRecords.xlsx')
    sheet = wb['Sheet1']
    
    lastCol = sheet.max_column
    latestMonth = sheet.cell(row=1, column=lastCol).value
    latestMonth = datetime.datetime.strftime(latestMonth, '%b %Y')  # added for LibreOffice 6.0.3.2
    
    # Check each member's payment status.
    unpaidMembers = {}
    for r in range(2, sheet.max_row + 1):
        payment = sheet.cell(row=r, column=lastCol).value
        if payment != 'paid':
            name = sheet.cell(row=r, column=1).value
            email = sheet.cell(row=r, column=2).value
            unpaidMembers[name] = email
    
    # Log in to email account.
    with open('smtp_info') as config:
        # smtp_cfg = [email, password, smtp server, port]
        myEmail, password, server, port = config.read().splitlines()
    
    smtpObj = smtplib.SMTP_SSL(server, port)  # Using port 465
    smtpObj.ehlo()
    smtpObj.login(myEmail, password)
    
    # Send out reminder emails.
    for name, email in unpaidMembers.items():
        body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. " \
               "Please make this payment as soon as possible. Thank you!" % (latestMonth, name, latestMonth)
        print(f'Sending email to {email}...')
        sendmailStatus = smtpObj.sendmail(myEmail, email, body)
    
        if sendmailStatus != {}:
            print(f'There was a problem sending email to {email}: {sendmailStatus}')
    smtpObj.quit()


if __name__ == '__main__':
    main()

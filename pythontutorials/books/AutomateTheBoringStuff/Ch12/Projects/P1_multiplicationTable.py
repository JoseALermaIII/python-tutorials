"""Multiplication table

Create a program multiplicationTable.py that takes a number N from the
command line and creates an N×N multiplication table in an Excel spreadsheet.
Row 1 and column A should be used for labels and should be in bold.

"""


def main():
    import sys
    import openpyxl
    from openpyxl.styles import Font

    # Get argument from commandline
    size = int(''.join(sys.argv[1]))

    # Build table
    wb = openpyxl.Workbook()
    sheet = wb.active
    row = []
    for i in range(1, size + 1):
        row.append(i)
    column = row

    # Make labels
    for element in row:
        sheet.cell(row=1, column=element + 1).value = element
        sheet.cell(row=element + 1, column=1).value = element
    col = sheet.column_dimensions['A']
    col.font = Font(bold=True)
    ro = sheet.row_dimensions[1]
    ro.font = Font(bold=True)

    # Calculate table values
    for element in row:
        for element2 in column:
            sheet.cell(row=element + 1, column=element2 + 1).value = element * element2

    # Save table
    wb.save("multTable.xlsx")


if __name__ == '__main__':
    main()

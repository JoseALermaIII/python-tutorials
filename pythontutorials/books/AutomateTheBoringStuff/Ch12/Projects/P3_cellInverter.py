"""Cell inverter

Write a program to invert the row and column of the cells in the spreadsheet.
For example, the value at row 5, column 3 will be at row 3, column 5
(and vice versa). This should be done for all cells in the spreadsheet.

Note:
    Gets full file path from commandline arguments.

"""


def main():
    import sys
    import openpyxl

    # Get argument from commandline
    file = ''.join(sys.argv[1])

    # Open workbooks
    readwb = openpyxl.load_workbook(file)
    readsheet = readwb.active

    writewb = openpyxl.Workbook()
    writesheet = writewb.active

    # Read readwb and transpose into writewb
    for rowNum in range(1, readsheet.max_row + 1):
        for colNum in range(1, readsheet.max_column + 1):
                # Invert columns and rows
                writesheet.cell(row=colNum, column=rowNum).value = readsheet.cell(row=rowNum, column=colNum).value

    # Save workbook
    writewb.save("updated" + file)


if __name__ == '__main__':
    main()

"""Blank row inserter

Create a program blankRowInserter.py that takes two integers and a filename
string as command line arguments. Let’s call the first integer N and the second
integer M. Starting at row N, the program should insert M blank rows into the
spreadsheet.

"""


def main():
    import sys
    import openpyxl

    # Get arguments from commandline
    N = int(''.join(sys.argv[1]))
    M = int(''.join(sys.argv[2]))
    file = ''.join(sys.argv[3])

    # Open workbooks
    readwb = openpyxl.load_workbook(file)
    readsheet = readwb.active

    writewb = openpyxl.Workbook()
    writesheet = writewb.active

    # Read readwb and transpose into writewb
    for rowNum in range(1, readsheet.max_row + 1):
        for colNum in range(1, readsheet.max_column + 1):
                if rowNum < N:
                    writesheet.cell(row=rowNum, column=colNum).value = readsheet.cell(row=rowNum, column=colNum).value
                else:
                    # Insert blank lines
                    writesheet.cell(row=(rowNum + M), column=colNum).value = readsheet.cell(row=rowNum, column=colNum).value

    # Save workbook
    writewb.save("updated" + file)


if __name__ == '__main__':
    main()

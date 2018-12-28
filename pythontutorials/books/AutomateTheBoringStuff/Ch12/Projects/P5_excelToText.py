"""Excel to text

Write a program that performs the tasks of the previous program in reverse
order: The program should open a spreadsheet and write the cells of column A
into one text file, the cells of column B into another text file, and so on.

Note:
    Default output folder is ``./p5files/``.

"""


def main():
    import openpyxl
    import os

    # Make directory
    FOLDER = "./p5files/"
    os.makedirs(FOLDER, exist_ok=True)

    # Open workbook
    wb = openpyxl.load_workbook("example.xlsx")
    sheet = wb.active

    # Open file
    index = 1
    while index <= sheet.max_column:
        fileObj = open(FOLDER + "text" + str(index) + ".txt", 'w')
        for rowNum in range(1, sheet.max_row + 1):
            # Transpose column into relevant text file
            fileObj.write(str(sheet.cell(row=rowNum, column=index).value) + '\n')
        index += 1
        fileObj.close()


if __name__ == '__main__':
    main()

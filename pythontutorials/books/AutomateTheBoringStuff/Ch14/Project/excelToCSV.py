"""Excel to CSV

Using :py:mod:`openpyxl`, write a program that reads all the Excel files in the
current working directory and outputs them as CSV files.

A single Excel file might contain multiple sheets; you’ll have to create one CSV
file per sheet. The filenames of the CSV files should be
``<excel filename>_<sheet title>.csv``, where ``<excel filename>`` is the filename of the
Excel file without the file extension (for example, 'spam_data', not
'spam_data.xlsx') and ``<sheet title>`` is the string from the Worksheet object’s title
variable.

Notes:
    * Example Excel files can be downloaded from http://nostarch.com/automatestuff/
    * Default input folder is ``./excelSpreadsheets``.
    * Default output folder is ``./csvFiles``.

"""


def main():
    import openpyxl, os, csv

    FOLDER_IN = "./excelSpreadsheets"
    FOLDER_OUT = "./csvFiles"

    for excelFile in os.listdir(FOLDER_IN):
        # Skip non-xlsx files, load the workbook object.
        if excelFile.endswith(".xlsx"):
            if not FOLDER_IN.endswith('/'):
                excelFilepath = FOLDER_IN + '/' + excelFile
            else:
                excelFilepath = FOLDER_IN + excelFile
            wb = openpyxl.load_workbook(excelFilepath)
        else:
            continue
        for sheetName in wb.sheetnames:
            # Loop through every sheet in the workbook.
            sheet = wb[sheetName]

            # Create the CSV filename from the Excel filename and sheet title.
            if not FOLDER_OUT.endswith('/'):
                csvFilepath = FOLDER_OUT + '/' + excelFile[:-5] + '_' + sheet.title + ".csv"
            else:
                csvFilepath = FOLDER_OUT + excelFile[:-5] + '_' + sheet.title + ".csv"
            # Create the csv.writer object for this CSV file.
            csvFile = open(csvFilepath, "w", newline='')
            csvWriter = csv.writer(csvFile)
            # Loop through every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                rowData = []    # append each cell to this list
                # Loop through each cell in the row.
                for colNum in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData.
                    rowData.append(sheet.cell(row=rowNum, column=colNum).value)
                # Write the rowData list to the CSV file.
                csvWriter.writerow(rowData)
            csvFile.close()


if __name__ == '__main__':
    main()

# Using the openpyxl module, write a program that reads all the Excel files in the
# current working directory and outputs them as CSV files.
# A single Excel file might contain multiple sheets; you’ll have to create one CSV
# file per sheet. The filenames of the CSV files should be
# <excel filename>_<sheet title>.csv, where <excel filename> is the filename of the
# Excel file without the file extension (for example, 'spam_data', not
# 'spam_data.xlsx') and <sheet title> is the string from the Worksheet object’s title
# variable.
#
# Note:
# - Example Excel files can be downloaded from http://nostarch.com/automatestuff/

import openpyxl, os

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb['sheetName']

        # Create the CSV filename from the Excel filename and sheet title.
        # Create the csv.writer object for this CSV file.

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.

            # Write the rowData list to the CSV file.

        csvFile.close()
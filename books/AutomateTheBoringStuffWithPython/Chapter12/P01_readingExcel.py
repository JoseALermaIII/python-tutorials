# This program uses the OpenPyXL module to manipulate Excel documents

# Opening Excel Documents with OpenPyXL
import openpyxl
print(openpyxl.__version__)  # Using version 2.5.5

wb = openpyxl.load_workbook("example.xlsx")
print(type(wb))

# Getting Sheets from the Workbook
print(wb.sheetnames)
sheet = wb["Sheet3"]
print(sheet)
print(type(sheet))
print(sheet.title)
anotherSheet = wb.active
print(anotherSheet)

# Getting Cells from the Sheets
print(anotherSheet["A1"])
print(anotherSheet["A1"].value)
print(type(anotherSheet["A1"].value))
c = anotherSheet["B1"]
print(c.value)
print("Row " + str(c.row) + ", Column " + c.column + " is " + c.value)
print(anotherSheet["C1"].value)

print(anotherSheet.cell(row=1, column=2))
print(anotherSheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, anotherSheet.cell(row=i, column=2).value)

sheet = wb["Sheet1"]
print(sheet.max_row)
print(sheet.max_column)

# Converting Between Column Letters and Numbers
from openpyxl.utils import get_column_letter, column_index_from_string  # Imports should be at the top of the file
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string("AA"))

# Getting Rows and Columns from the Sheets
print(tuple(sheet["A1":"C3"]))
for rowOfCellObjects in sheet["A1":"C3"]:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print("--- END OF ROW ---")

print(sheet.columns)
for columnOfCellObjects in sheet.columns:
    for cellObj in columnOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print("--- END OF COLUMN ---")

print(list(sheet.columns)[1])
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)

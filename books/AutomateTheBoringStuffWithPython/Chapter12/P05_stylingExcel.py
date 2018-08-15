# This program uses the OpenPyXL module to manipulate Excel documents

import openpyxl
from openpyxl.styles import Font, NamedStyle

wb = openpyxl.Workbook()
sheet = wb["Sheet"]

# Setting the Font Style of Cells
italic24Font = NamedStyle(name="italic24Font")
italic24Font.font = Font(size=24, italic=True)
sheet["A1"].style = italic24Font
sheet["A1"] = "Hello world!"
wb.save("styled.xlsx")

# Font Objects
wb = openpyxl.Workbook()
sheet = wb["Sheet"]

fontObj1 = Font(name="Times New Roman", bold=True)
styleObj1 = NamedStyle(name="styleObj1")
styleObj1.font = fontObj1
sheet["A1"].style = styleObj1
sheet["A1"] = "Bold Times New Roman"

fontObj2 = Font(size=24, italic=True)
styleObj2 = NamedStyle(name="styleObj2")
styleObj2.font = fontObj2
sheet["B3"].style = styleObj2
sheet["B3"] = "24 pt Italic"
wb.save("styles.xlsx")

# Formulas
wb = openpyxl.Workbook()
sheet = wb.active
sheet["A1"] = 200
sheet["A2"] = 300
sheet["A3"] = "=SUM(A1:A2)"
wb.save("writeFormula.xlsx")

wbFormulas = openpyxl.load_workbook("writeFormula.xlsx")
sheet = wbFormulas.active
print(sheet["A3"].value)

wbDataOnly = openpyxl.load_workbook("writeFormula.xlsx", data_only=True)
sheet = wbDataOnly.active
print(sheet["A3"].value)  # Not working with LibreOffice 6.0.3.2

# Adjusting Rows and Columns
wb = openpyxl.Workbook()
sheet = wb.active
sheet["A1"] = "Tall row"
sheet["B2"] = "Wide column"
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save("dimensions.xlsx")

wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells("A1:D3")
sheet["A1"] = "Twelve cells merged together."
sheet.merge_cells("C5:D5")
sheet["C5"] = "Two merged cells."
wb.save("merged.xlsx")

wb = openpyxl.load_workbook("merged.xlsx")
sheet = wb.active
sheet.unmerge_cells("A1:D3")
sheet.unmerge_cells("C5:D5")
#wb.save("merged.xlsx")  # uncomment to see changes

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb.active
sheet.freeze_panes = "A2"
wb.save("freezeExample.xlsx")

# Charts
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):    # create some data in column A
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_row=1, min_col=1, max_row=10, max_col=1)

seriesObj = openpyxl.chart.Series(refObj, title="First Series")  # FIXME: Chart layout is wrong

chartObj = openpyxl.chart.BarChart()
chartObj.append(seriesObj)
chartObj.anchor = "B3"  # set the position
chartObj.width = 7.94  # set the size (in centimeters, where 1 cm = 37.8 pixels)
chartObj.height = 5.29

sheet.add_chart(chartObj)
wb.save("sampleChart.xlsx")

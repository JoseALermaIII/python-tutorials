# This program uses the OpenPyXL module to manipulate Excel documents

# Creating and Saving Excel Documents

import openpyxl
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
print(sheet.title)
sheet.title = "Spam Bacon Eggs Sheet"
print(wb.sheetnames)

wb = openpyxl.load_workbook("example.xlsx")
sheet = wb.active
sheet.title = "Spam Spam Spam"
wb.save("example_copy.xlsx")

# Creating and Removing Sheets
wb = openpyxl.Workbook()
print(wb.sheetnames)
wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet(index=0, title="First Sheet")
print(wb.sheetnames)
wb.create_sheet(index=2, title="Middle Sheet")
print(wb.sheetnames)

wb.remove(wb["Middle Sheet"])
wb.remove(wb["Sheet1"])
print(wb.sheetnames)

# Writing Values to Cells
sheet = wb.get_sheet_by_name("Sheet")
sheet["A1"] = "Hello world!"
print(sheet["A1"].value)
